"""
Report Generator
Creates PDF and Excel reports summarising cut/fill volume results.
"""

import io
import datetime
from typing import List

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from matplotlib.patches import Patch

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image as RLImage, HRFlowable,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from volume_calculator import VolumeResult


# ─────────────────────────────────────────────
# Matplotlib helper — render cut/fill heatmap
# ─────────────────────────────────────────────

def _render_heatmap_png(result: VolumeResult, width_px: int = 800) -> bytes:
    """Render the dz grid as a PNG and return raw bytes."""
    dz_arr = np.array(result.dz_grid, dtype=float)

    # Symmetric colour scale
    abs_max = np.nanmax(np.abs(dz_arr))
    if abs_max == 0:
        abs_max = 1.0

    cmap = mcolors.LinearSegmentedColormap.from_list(
        "cut_fill", ["#27AE60", "#FFFFFF", "#E74C3C"]
    )

    fig, ax = plt.subplots(figsize=(10, 8), dpi=100)
    im = ax.imshow(
        dz_arr,
        origin="lower",
        extent=[min(result.grid_x), max(result.grid_x),
                min(result.grid_y), max(result.grid_y)],
        cmap=cmap,
        vmin=-abs_max,
        vmax=abs_max,
        aspect="equal",
        interpolation="bilinear",
    )
    cbar = fig.colorbar(im, ax=ax, label="Δ Elevation (m)  [+ve = Cut, −ve = Fill]")
    ax.set_title(
        f"Cut / Fill Map\n{result.surface1_name}  vs  {result.surface2_name}",
        fontsize=13, fontweight="bold",
    )
    ax.set_xlabel("Easting (m)")
    ax.set_ylabel("Northing (m)")

    legend_elements = [
        Patch(facecolor="#E74C3C", label="Cut (surface1 above surface2)"),
        Patch(facecolor="#27AE60", label="Fill (surface2 above surface1)"),
    ]
    ax.legend(handles=legend_elements, loc="upper right", fontsize=8)
    ax.tick_params(axis="both", labelsize=8)

    buf = io.BytesIO()
    plt.savefig(buf, format="png", bbox_inches="tight", dpi=100)
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
# PDF Report
# ─────────────────────────────────────────────

def generate_pdf(results: List[VolumeResult], project_name: str = "Earthworks Project") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title2", parent=styles["Title"], fontSize=20, spaceAfter=6
    )
    heading_style = ParagraphStyle(
        "Heading2b", parent=styles["Heading2"], fontSize=13, spaceAfter=4
    )
    body_style = styles["BodyText"]
    caption_style = ParagraphStyle(
        "Caption", parent=styles["BodyText"], fontSize=8,
        textColor=colors.gray, alignment=TA_CENTER,
    )

    story = []

    # ── Cover / Header ──
    story.append(Paragraph(project_name, title_style))
    story.append(Paragraph("Cut / Fill Volume Report", styles["Heading1"]))
    now = datetime.datetime.now().strftime("%d %B %Y  %H:%M")
    story.append(Paragraph(f"Generated: {now}", body_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2C3E50")))
    story.append(Spacer(1, 0.4 * cm))

    # ── Summary table ──
    story.append(Paragraph("Volume Summary", heading_style))
    story.append(Spacer(1, 0.2 * cm))

    table_data = [
        ["Surface 1 (Existing)", "Surface 2 (Reference)",
         "Cut (m³)", "Fill (m³)", "Net (m³)"],
    ]
    for r in results:
        table_data.append([
            r.surface1_name,
            r.surface2_name,
            f"{r.cut_volume:,.3f}",
            f"{r.fill_volume:,.3f}",
            f"{r.net_volume:+,.3f}",
        ])

    col_widths = [4.2 * cm, 4.2 * cm, 3 * cm, 3 * cm, 3 * cm]
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#F2F2F2"), colors.white]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ALIGN",      (2, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.6 * cm))

    # ── Per-result section ──
    for idx, r in enumerate(results):
        story.append(Paragraph(
            f"Calculation {idx + 1}: {r.surface1_name}  vs  {r.surface2_name}",
            heading_style,
        ))

        details = [
            ["Parameter", "Value"],
            ["Surface 1 (Existing / Base)", r.surface1_name],
            ["Surface 2 (Reference / Design)", r.surface2_name],
            ["Grid Resolution", f"{r.grid_resolution:.3f} m"],
            ["Grid Size", f"{len(r.grid_x)} × {len(r.grid_y)} cells"],
            ["Cut Volume", f"{r.cut_volume:,.3f} m³"],
            ["Fill Volume", f"{r.fill_volume:,.3f} m³"],
            ["Net Volume (Cut − Fill)", f"{r.net_volume:+,.3f} m³"],
        ]
        dt = Table(details, colWidths=[6.5 * cm, 10.5 * cm])
        dt.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#3498DB")),
            ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
            ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME",    (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#EBF5FB"), colors.white]),
            ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(dt)
        story.append(Spacer(1, 0.4 * cm))

        # Heatmap image
        try:
            png_bytes = _render_heatmap_png(r)
            img_buf = io.BytesIO(png_bytes)
            img = RLImage(img_buf, width=16 * cm, height=12 * cm)
            story.append(img)
            story.append(Paragraph(
                f"Figure {idx + 1}: Cut / Fill map — red = cut, green = fill",
                caption_style,
            ))
        except Exception:
            story.append(Paragraph("[Map image unavailable]", body_style))

        story.append(Spacer(1, 0.8 * cm))

    # ── Footer note ──
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.gray))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        "Note: Volumes calculated using a grid-based interpolation method. "
        "Cut = existing ground above reference; Fill = reference above existing ground.",
        ParagraphStyle("Note", parent=body_style, fontSize=8, textColor=colors.gray),
    ))

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
# Excel Report
# ─────────────────────────────────────────────

def generate_excel(results: List[VolumeResult], project_name: str = "Earthworks Project") -> bytes:
    wb = openpyxl.Workbook()

    # ── Summary sheet ──
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font  = Font(bold=True, size=14)
    sub_font    = Font(bold=True, size=11)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _cell(ws, row, col, value, font=None, fill=None, align=None, number_format=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:   c.font = font
        if fill:   c.fill = fill
        if align:  c.alignment = align
        if number_format: c.number_format = number_format
        c.border = border
        return c

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = project_name
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Cut / Fill Volume Report — Generated {datetime.datetime.now().strftime('%d %B %Y')}"
    ws["A2"].alignment = Alignment(horizontal="center")
    ws["A2"].font = Font(size=10, italic=True, color="666666")

    # Summary table headers
    headers = ["Surface 1", "Surface 2", "Cut Volume (m³)", "Fill Volume (m³)", "Net Volume (m³)", "Grid Res (m)"]
    for ci, h in enumerate(headers, start=1):
        _cell(ws, 4, ci, h, font=header_font, fill=header_fill,
              align=Alignment(horizontal="center", wrap_text=True))

    alt_fill_a = PatternFill("solid", fgColor="EBF5FB")
    alt_fill_b = PatternFill("solid", fgColor="FFFFFF")

    for ri, r in enumerate(results):
        row = 5 + ri
        fill = alt_fill_a if ri % 2 == 0 else alt_fill_b
        _cell(ws, row, 1, r.surface1_name, fill=fill)
        _cell(ws, row, 2, r.surface2_name, fill=fill)
        _cell(ws, row, 3, round(r.cut_volume, 3),  fill=fill, number_format='#,##0.000')
        _cell(ws, row, 4, round(r.fill_volume, 3), fill=fill, number_format='#,##0.000')
        _cell(ws, row, 5, round(r.net_volume, 3),  fill=fill, number_format='+#,##0.000;-#,##0.000')
        _cell(ws, row, 6, round(r.grid_resolution, 3), fill=fill)

    col_widths = [25, 25, 18, 18, 18, 12]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[4].height = 30

    # ── Per-result detail sheets ──
    for idx, r in enumerate(results):
        sheet_name = f"Detail_{idx + 1}"[:31]
        ds = wb.create_sheet(title=sheet_name)

        ds.merge_cells("A1:C1")
        ds["A1"] = f"{r.surface1_name}  vs  {r.surface2_name}"
        ds["A1"].font = title_font
        ds["A1"].alignment = Alignment(horizontal="left")

        details = [
            ("Surface 1 (Existing)", r.surface1_name),
            ("Surface 2 (Reference)", r.surface2_name),
            ("Grid Resolution (m)", round(r.grid_resolution, 4)),
            ("Grid Columns", len(r.grid_x)),
            ("Grid Rows", len(r.grid_y)),
            ("Cut Volume (m³)", round(r.cut_volume, 3)),
            ("Fill Volume (m³)", round(r.fill_volume, 3)),
            ("Net Volume (m³)", round(r.net_volume, 3)),
        ]

        for ri, (k, v) in enumerate(details, start=3):
            ds.cell(row=ri, column=1, value=k).font = Font(bold=True)
            ds.cell(row=ri, column=2, value=v)

        ds.column_dimensions["A"].width = 30
        ds.column_dimensions["B"].width = 25

        # Raw dz grid sheet
        grid_sheet_name = f"Grid_{idx + 1}"[:31]
        gs = wb.create_sheet(title=grid_sheet_name)
        gs.cell(row=1, column=1, value="Δz Grid (rows = Northing ↑, cols = Easting →)")
        gs.cell(row=1, column=1).font = Font(bold=True)

        # Write X axis header
        for ci, xv in enumerate(r.grid_x, start=2):
            gs.cell(row=2, column=ci, value=round(xv, 2))

        dz_arr = r.dz_grid
        for ri, (yv, row_vals) in enumerate(zip(r.grid_y, dz_arr), start=3):
            gs.cell(row=ri, column=1, value=round(yv, 2))
            for ci, val in enumerate(row_vals, start=2):
                c = gs.cell(row=ri, column=ci, value=round(val, 4) if val is not None else "")
                if val is not None:
                    if val > 0:
                        c.fill = PatternFill("solid", fgColor="FFB3AE")   # red tint = cut
                    elif val < 0:
                        c.fill = PatternFill("solid", fgColor="B3FFB8")   # green tint = fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
